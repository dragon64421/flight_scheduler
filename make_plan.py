import openpyxl
import random
import sys
import numpy

def load_crew_data(filename): #엑셀에서 crew data 불러오기
    crew_data = openpyxl.load_workbook(filename)
    crew_data_sheet_o = crew_data.worksheets[0]
    crew_data_sheet_nco = crew_data.worksheets[1]

    index = 1
    crew_data_o = []
    while crew_data_sheet_o.cell(row=index, column=1).value != '0999':
        new_crew_data_o = []
        for i in range(1,5):
            new_crew_data_o.append(str(crew_data_sheet_o.cell(row=index, column=i).value))
        new_crew_data_o.append(0)
        crew_data_o.append(new_crew_data_o)
        index += 1

    index = 1
    crew_data_nco = []
    while crew_data_sheet_nco.cell(row=index, column=1).value != '1999':
        new_crew_data_nco = []
        for i in range(1,5):
            new_crew_data_nco.append(str(crew_data_sheet_nco.cell(row=index, column=i).value))
        new_crew_data_nco.append(0)
        crew_data_nco.append(new_crew_data_nco)
        index += 1

    return crew_data_o + crew_data_nco

def load_mission_info(filename): #엑셀에서 mission info 불러오기
    mission_info = openpyxl.load_workbook(filename)
    mission_info_sheet = mission_info.worksheets[2]

    index = 2
    mission = []
    while mission_info_sheet.cell(row=index, column=1).value != None:
        new_mission =[]
        for i in range(1,11):
            new_mission.append(str(mission_info_sheet.cell(row=index, column=i).value))
        excluded_crew = []
        for i in range(11,22):
            if mission_info_sheet.cell(row=index, column=i).value != None:
                excluded_crew = excluded_crew + (str(mission_info_sheet.cell(row=index, column=i).value)).split(',')
        new_mission.append(excluded_crew)
        for i in range(11,22):
            if mission_info_sheet.cell(row=index, column=i).value != None:
                new_mission.append((str(mission_info_sheet.cell(row=index, column=i).value)).split(','))
            else:
                new_mission.append(['-'])
        mission.append(new_mission)
        index += 1

    return mission

def sort_qualification(available_crew_data, quality): #(결원을 제외한) crew data에서 선택한 직급의 인원을 추출
    sorted = []
    for i in range(0, len(available_crew_data)):
        if available_crew_data[i][2] == str(quality):
            sorted.append(available_crew_data[i])
    return sorted

def exclude_crew(crew_data_up, mission):
    excluded_crew = mission[10]
    crew_name = []
    for i in range(0, len(crew_data_up)):
        crew_name.append(crew_data_up[i][1])
    excluded_crew_index = []
    for i in range(0, len(excluded_crew)):
        excluded_crew_index.append(crew_name.index(excluded_crew[i]))
    excluded_crew_index.sort(reverse=True)
    excluded_crew_data = []
    for i in excluded_crew_index:
        excluded_crew_data.append(crew_data_up[i])
    for i in range(0, len(excluded_crew_data)):
        crew_data_up.remove(excluded_crew_data[i])
    return crew_data_up

def pick_a_crew(qual_sorted_crew_data):
    days_1 = []
    for i in range(0, len(qual_sorted_crew_data)):
        days_1.append(int(qual_sorted_crew_data[i][3])+0.000001)
    mean_1 = numpy.mean(days_1)
    days_1 = [days_1[i] - mean_1 + 0.0000001 for i in range(len(days_1))]
    std_1 = numpy.std(days_1)
    if std_1 != 0:
        days_1 = [days_1[i]/std_1 for i in range(len(days_1))]
    else:
        days_1 = [0 for i in range(len(days_1))]
    days_2 = []
    for i in range(0, len(qual_sorted_crew_data)):
        days_2.append(int(qual_sorted_crew_data[i][4])+0.0000001)
    mean_2 = numpy.mean(days_2)
    days_2 = [days_2[i] - mean_2 for i in range(len(days_2))]
    std_2 = numpy.std(days_2)
    if std_2 != 0:
        days_2 = [days_2[i] / std_2 for i in range(len(days_2))]
    else:
        days_2 = [0 for i in range(len(days_2))]
    days = [days_1[i] - days_2[i] for i in range(len(days_1))]
    index = [i for i, value in enumerate(days) if value == max(days)]
    if len(index) != 1:
        x = random.randint(10, 30)
        index = [index[x % len(index)]]
    return qual_sorted_crew_data[index[0]] #crew data

def counting(crew_data_updated, picked_crew_names):
    crew_name = []
    for i in range(0, len(crew_data_updated)):
        crew_name.append(crew_data_updated[i][1])
    picked_crew_index = []
    for i in range(0, len(picked_crew_names)):
        picked_crew_index.append(crew_name.index(picked_crew_names[i]))
    for i in range(0, len(crew_data_updated)):
        crew_data_updated[i][3] = str(int(crew_data_updated[i][3])+1)
    for i in range(0, len(picked_crew_index)):
        crew_data_updated[picked_crew_index[i]][3] = '0'
    for i in range(0, len(picked_crew_index)):
        count = crew_data_updated[picked_crew_index[i]][4]
        crew_data_updated[picked_crew_index[i]][4] = count + 1
    return crew_data_updated

def make_mission_plan(one_mission_info, crew_data_updated):
    mission_crew = []
    crew_data_up = crew_data_updated[:]

    pre_picked_crew_names = []
    for i in range(14, 22):
        pre_picked_crew_names.append(str(one_mission_info[i][0]))

    available_crew_data = exclude_crew(crew_data_up, one_mission_info)[:]
    index = 2
    for qual in ['MC', 'ASO', 'IDO', 'WAO', 'WD', 'SO1', 'SO2', 'SO3']:
        crew_name = []
        for i in range(0, len(crew_data_updated)):
            crew_name.append(crew_data_updated[i][1])

        picked_crew_index = []
        if pre_picked_crew_names[index-2] != '-':
            picked_crew_index.append(crew_name.index(pre_picked_crew_names[index-2]))
            mission_crew.append(crew_data_updated[picked_crew_index[0]][0:3])

        qual_sorted_from_avail_crew_data = sort_qualification(available_crew_data, qual)[:]
        for _ in range(0, int(one_mission_info[index])):
            picked_crew = pick_a_crew(qual_sorted_from_avail_crew_data)
            mission_crew.append(picked_crew[0:3])
            qual_sorted_from_avail_crew_data.remove(picked_crew)
        index += 1

    picked_crew_names = []
    for i in range(0, len(mission_crew)):
        picked_crew_names.append(mission_crew[i][1])
    crew_data_updated = counting(crew_data_updated, picked_crew_names)

    return [mission_crew, crew_data_updated]

def write_schedule(schedule, crew_data_updated, mission_info):
    flight_schedule = openpyxl.load_workbook("flight_schedule_f.xlsx")
    flight_schedule_sheet = flight_schedule.worksheets[0]

    for index in range(0, len(mission_info)):
        flight_schedule_sheet.cell(row=3*index+2, column=2, value=mission_info[index][0])
        flight_schedule_sheet.cell(row=3*index+2, column=3, value=mission_info[index][1])
        for i in range(11, 14):
            if mission_info[index][i] != ['-']:
                names = ''
                for k in range(0, len(mission_info[index][i])):
                    names = names + ' ' + str(mission_info[index][i][k])
                    flight_schedule_sheet.cell(row=3*index+2, column=5+i, value=names)
        quals = []
        for g in range(0, len(schedule[index])):
            quals.append(str(schedule[index][g][2]))
        rowcount = 0
        for qual in ['MC', 'ASO', 'IDO', 'WAO', 'WD']:
            where = [h for h, value in enumerate(quals) if value == qual]
            for t in range(0, len(where)):
                if rowcount > 5:
                    flight_schedule_sheet.cell(row=3 * index + 4, column=2 * (rowcount - 6) + 4,
                                               value=schedule[index][rowcount][2])
                    flight_schedule_sheet.cell(row=3 * index + 4, column=2 * (rowcount - 6) + 5,
                                               value=schedule[index][rowcount][1])
                    rowcount += 1
                elif rowcount > 2:
                    flight_schedule_sheet.cell(row=3 * index + 3, column=2 * (rowcount - 3) + 4,
                                               value=schedule[index][rowcount][2])
                    flight_schedule_sheet.cell(row=3 * index + 3, column=2 * (rowcount - 3) + 5,
                                               value=schedule[index][rowcount][1])
                    rowcount += 1
                else:
                    flight_schedule_sheet.cell(row=3 * index + 2, column=2 * rowcount + 4,
                                               value=schedule[index][rowcount][2])
                    flight_schedule_sheet.cell(row=3 * index + 2, column=2 * rowcount + 5,
                                               value=schedule[index][rowcount][1])
                    rowcount += 1
        crewcount = rowcount
        rowcount = 0
        for qual in ['SO1', 'SO2', 'SO3']:
            where = [h for h, value in enumerate(quals) if value == qual]
            for t in range(0, len(where)):
                if rowcount > 5:
                    flight_schedule_sheet.cell(row=3 * index + 4, column=2 * (rowcount - 6) + 10,
                                               value=schedule[index][crewcount][2])
                    flight_schedule_sheet.cell(row=3 * index + 4, column=2 * (rowcount - 6) + 11,
                                               value=schedule[index][crewcount][1])
                    rowcount += 1
                    crewcount += 1
                elif rowcount > 2:
                    flight_schedule_sheet.cell(row=3 * index + 3, column=2 * (rowcount - 3) + 10,
                                               value=schedule[index][crewcount][2])
                    flight_schedule_sheet.cell(row=3 * index + 3, column=2 * (rowcount - 3) + 11,
                                               value=schedule[index][crewcount][1])
                    rowcount += 1
                    crewcount += 1
                else:
                    flight_schedule_sheet.cell(row=3 * index + 2, column=2 * rowcount + 10,
                                               value=schedule[index][crewcount][2])
                    flight_schedule_sheet.cell(row=3 * index + 2, column=2 * rowcount + 11,
                                               value=schedule[index][crewcount][1])
                    rowcount += 1
                    crewcount += 1

    index = 0
    while crew_data_updated[index][2] != 'SO1':
        flight_schedule_sheet.cell(row=index + 2, column=21, value=str(crew_data_updated[index][0]))
        flight_schedule_sheet.cell(row=index + 2, column=22, value=str(crew_data_updated[index][1]))
        flight_schedule_sheet.cell(row=index + 2, column=23, value=str(crew_data_updated[index][2]))
        flight_schedule_sheet.cell(row=index + 2, column=24, value=str(crew_data_updated[index][4]))
        index += 1
    rowcount = 0
    while index != len(crew_data_updated):
        flight_schedule_sheet.cell(row=rowcount + 2, column=26, value=str(crew_data_updated[index][0]))
        flight_schedule_sheet.cell(row=rowcount + 2, column=27, value=str(crew_data_updated[index][1]))
        flight_schedule_sheet.cell(row=rowcount + 2, column=28, value=str(crew_data_updated[index][2]))
        flight_schedule_sheet.cell(row=rowcount + 2, column=29, value=str(crew_data_updated[index][4]))
        index += 1
        rowcount += 1

    flight_schedule.save('schedule(backup).xlsx')

    crew_data = openpyxl.load_workbook('crew_data.xlsx')
    crew_data_sheet_o = crew_data.worksheets[0]
    crew_data_sheet_nco = crew_data.worksheets[1]
    index = 0
    while crew_data_updated[index][2] != 'SO1':
        crew_data_sheet_o.cell(row=index + 1, column=1, value=str(crew_data_updated[index][0]))
        crew_data_sheet_o.cell(row=index + 1, column=2, value=str(crew_data_updated[index][1]))
        crew_data_sheet_o.cell(row=index + 1, column=3, value=str(crew_data_updated[index][2]))
        crew_data_sheet_o.cell(row=index + 1, column=4, value=str(crew_data_updated[index][3]))
        index += 1
    rowcount = 0
    while index != len(crew_data_updated):
        crew_data_sheet_nco.cell(row=rowcount + 1, column=1, value=str(crew_data_updated[index][0]))
        crew_data_sheet_nco.cell(row=rowcount + 1, column=2, value=str(crew_data_updated[index][1]))
        crew_data_sheet_nco.cell(row=rowcount + 1, column=3, value=str(crew_data_updated[index][2]))
        crew_data_sheet_nco.cell(row=rowcount + 1, column=4, value=str(crew_data_updated[index][3]))
        index += 1
        rowcount += 1
    crew_data.save('updated_crew_data.xlsx')

    return

#def crew_data_days_update(crew_data_updated):