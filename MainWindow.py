from PyQt5.QtWidgets import *
from PyQt5.uic import *
from PyQt5.QtCore import pyqtSlot, Qt
from PyQt5 import QtGui
import openpyxl
from make_plan import *

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        loadUi("flight_scheduler.ui", self)
        self.Initialize()

    def Initialize(self):
        self.setWindowTitle("Flight Scheduler")
        self.add_officer_button.clicked.connect(self.add_officer)
        self.add_nco_button.clicked.connect(self.add_nco)
        self.save_button.clicked.connect(self.save_crew_data)
        self.make_plan_button.clicked.connect(self.make_plan)
        #self.save_plan_button.clicked.connect(self.change_plan)
        self.check_mode = 0
        self.flight_schedule_table.itemChanged.connect(self.check_schedule)
        self.load_crew_data()

        self.load_mission_button.clicked.connect(self.load_mission_info)
        self.save_mission_button.clicked.connect(self.save_mission_info)
        index = 0
        for col in [30, 85, 30, 30, 30, 30, 30, 30, 30, 30, 80, 80, 80, 45, 45, 45, 45, 45, 45, 45, 45]:
            self.mission_info_table.setColumnWidth(index, col)
            index += 1
        index = 0
        for col in [30, 85, 35, 45, 35, 45, 35, 45, 35, 45, 35, 45, 35, 45, 80, 80, 80]:
            self.flight_schedule_table.setColumnWidth(index, col)
            index += 1
        self.flight_schedule_table.setRowCount(42)
        for i in range(0, 14):
            self.flight_schedule_table.setSpan(3 * i, 0, 3, 1)
            self.flight_schedule_table.setSpan(3 * i, 1, 3, 1)
            self.flight_schedule_table.setSpan(3 * i, 14, 3, 1)
            self.flight_schedule_table.setSpan(3 * i, 15, 3, 1)
            self.flight_schedule_table.setSpan(3 * i, 16, 3, 1)

    def load_crew_data(self):
        crew_data = openpyxl.load_workbook('crew_data.xlsx')
        crew_data_sheet_o = crew_data.worksheets[0]
        crew_data_sheet_nco = crew_data.worksheets[1]

        index = 1
        crew_data_o = []
        while crew_data_sheet_o.cell(row=index, column=1).value != '0999':
            new_crew_data_o = []
            for i in range(1, 5):
                new_crew_data_o.append(str(crew_data_sheet_o.cell(row=index, column=i).value))
            new_crew_data_o.append(0)
            crew_data_o.append(new_crew_data_o)
            index += 1

        index = 1
        crew_data_nco = []
        while crew_data_sheet_nco.cell(row=index, column=1).value != '1999':
            new_crew_data_nco = []
            for i in range(1, 5):
                new_crew_data_nco.append(str(crew_data_sheet_nco.cell(row=index, column=i).value))
            new_crew_data_nco.append(0)
            crew_data_nco.append(new_crew_data_nco)
            index += 1

        self.crew_data_o.setRowCount(len(crew_data_o))
        self.crew_data_nco.setRowCount(len(crew_data_nco))

        for i in range(0, len(crew_data_o)):
            item = QTableWidgetItem(str(crew_data_o[i][0]))
            item.setTextAlignment(Qt.AlignCenter)
            self.crew_data_o.setItem(i, 0, item)
            item = QTableWidgetItem(str(crew_data_o[i][1]))
            item.setTextAlignment(Qt.AlignCenter)
            self.crew_data_o.setItem(i, 1, item)
            item = QTableWidgetItem(str(crew_data_o[i][2]))
            item.setTextAlignment(Qt.AlignCenter)
            self.crew_data_o.setItem(i, 2, item)
            item = QTableWidgetItem(str(crew_data_o[i][3]))
            item.setTextAlignment(Qt.AlignCenter)
            self.crew_data_o.setItem(i, 3, item)

        for i in range(0, len(crew_data_nco)):
            item = QTableWidgetItem(str(crew_data_nco[i][0]))
            item.setTextAlignment(Qt.AlignCenter)
            self.crew_data_nco.setItem(i, 0, item)
            item = QTableWidgetItem(str(crew_data_nco[i][1]))
            item.setTextAlignment(Qt.AlignCenter)
            self.crew_data_nco.setItem(i, 1, item)
            item = QTableWidgetItem(str(crew_data_nco[i][2]))
            item.setTextAlignment(Qt.AlignCenter)
            self.crew_data_nco.setItem(i, 2, item)
            item = QTableWidgetItem(str(crew_data_nco[i][3]))
            item.setTextAlignment(Qt.AlignCenter)
            self.crew_data_nco.setItem(i, 3, item)

    def add_officer(self):
        rowcount = self.crew_data_o.rowCount()
        self.crew_data_o.setRowCount(rowcount + 1)
        for i in range(0, 4):
            item = QTableWidgetItem("")
            item.setTextAlignment(Qt.AlignCenter)
            self.crew_data_o.setItem(rowcount, i, item)

    def add_nco(self):
        rowcount = self.crew_data_nco.rowCount()
        self.crew_data_nco.setRowCount(rowcount + 1)
        for i in range(0, 4):
            item = QTableWidgetItem("")
            item.setTextAlignment(Qt.AlignCenter)
            self.crew_data_nco.setItem(rowcount, i, item)

    def save_crew_data(self):
        new_crew_data_o = []
        for i in range(0, self.crew_data_o.rowCount()):
            if self.crew_data_o.item(i,0) != None: #빈칸인데도 들어가는 버그 있는거 같음
                one_crew = []
                for k in range(0, 4):
                    one_crew.append(self.crew_data_o.item(i, k).text())
                new_crew_data_o.append(one_crew)
        new_crew_data_nco = []
        for i in range(0, self.crew_data_nco.rowCount()):
            if self.crew_data_nco.item(i,0) != None:
                one_crew = []
                for k in range(0, 4):
                    one_crew.append(self.crew_data_nco.item(i, k).text())
                new_crew_data_nco.append(one_crew)

        old_crew_data = openpyxl.load_workbook('crew_data.xlsx')
        old_crew_data.remove_sheet(old_crew_data["officer"])
        old_crew_data.remove_sheet(old_crew_data["nco"])
        old_crew_data.create_sheet('officer', 0)
        old_crew_data.create_sheet('nco', 1)
        officer_sheet = old_crew_data['officer']
        nco_sheet = old_crew_data['nco']

        for i in range(0, len(new_crew_data_o)):
            officer_sheet.cell(row=i + 1, column=1, value=str(new_crew_data_o[i][0]))
            officer_sheet.cell(row=i + 1, column=2, value=str(new_crew_data_o[i][1]))
            officer_sheet.cell(row=i + 1, column=3, value=str(new_crew_data_o[i][2]))
            officer_sheet.cell(row=i + 1, column=4, value=str(new_crew_data_o[i][3]))
        officer_sheet.cell(row=len(new_crew_data_o) + 1, column=1, value='0999')
        officer_sheet.cell(row=len(new_crew_data_o) + 1, column=2, value='###')
        officer_sheet.cell(row=len(new_crew_data_o) + 1, column=3, value='###')
        officer_sheet.cell(row=len(new_crew_data_o) + 1, column=4, value='###')

        for i in range(0, len(new_crew_data_nco)):
            nco_sheet.cell(row=i + 1, column=1, value=str(new_crew_data_nco[i][0]))
            nco_sheet.cell(row=i + 1, column=2, value=str(new_crew_data_nco[i][1]))
            nco_sheet.cell(row=i + 1, column=3, value=str(new_crew_data_nco[i][2]))
            nco_sheet.cell(row=i + 1, column=4, value=str(new_crew_data_nco[i][3]))
        nco_sheet.cell(row=len(new_crew_data_nco) + 1, column=1, value='1999')
        nco_sheet.cell(row=len(new_crew_data_nco) + 1, column=2, value='###')
        nco_sheet.cell(row=len(new_crew_data_nco) + 1, column=3, value='###')
        nco_sheet.cell(row=len(new_crew_data_nco) + 1, column=4, value='###')

        old_crew_data.save('crew_data.xlsx')

    def make_plan(self): #make_plan.py랑 연결됨
        mission_info = load_mission_info('crew_data.xlsx')
        crew_data = load_crew_data('crew_data.xlsx')
        self.check_mode = 0

        schedule = []
        crew_data_updated = crew_data
        for i in range(0, len(mission_info)):
            [mission_crew, crew_data_updated] = make_mission_plan(mission_info[i], crew_data_updated)
            schedule.append(mission_crew)

        write_schedule(schedule, crew_data_updated, mission_info)  # backup on excel

        while self.flight_crew.rowCount() != 0:
            self.flight_crew.removeRow(0)
        self.flight_crew.setRowCount(len(crew_data_updated))
        self.flight_crew.setColumnWidth(1, 35)
        self.flight_crew.setColumnWidth(2, 40)
        for i in range(0, len(crew_data_updated)):
            index = 0
            for j in [1, 2, 4]:
                item = QTableWidgetItem(str(crew_data_updated[i][j]))
                item.setTextAlignment(Qt.AlignCenter)
                self.flight_crew.setItem(i, index, item)
                index += 1

        while self.flight_schedule_table.rowCount() != 0:
            self.flight_schedule_table.removeRow(0)
        self.flight_schedule_table.setRowCount(3*len(schedule))
        for i in range(0, len(mission_info)):
            self.flight_schedule_table.setSpan(3 * i, 0, 3, 1)
            self.flight_schedule_table.setSpan(3 * i, 1, 3, 1)
            self.flight_schedule_table.setSpan(3 * i, 14, 3, 1)
            self.flight_schedule_table.setSpan(3 * i, 15, 3, 1)
            self.flight_schedule_table.setSpan(3 * i, 16, 3, 1)
            for j in [0, 1]:
                item = QTableWidgetItem(str(mission_info[i][j]))
                item.setTextAlignment(Qt.AlignCenter)
                self.flight_schedule_table.setItem(3*i, j, item)
            for j in [11, 12, 13]:
                if mission_info[i][j] != ['-']:
                    names = str(mission_info[i][j][0])
                    if len(mission_info[i][j]) != 1:
                        for k in range( 1, len(mission_info[i][j])):
                            names = names + ', ' +str(mission_info[i][j][k])
                    item = QTableWidgetItem(names)
                    item.setTextAlignment(Qt.AlignCenter)
                else:
                    item = QTableWidgetItem('')
                    item.setTextAlignment(Qt.AlignCenter)
                self.flight_schedule_table.setItem(3 * i, j + 3, item)
            quals = []
            for j in range(0, len(schedule[i])):
                quals.append(str(schedule[i][j][2]))
            o_count = 0
            for qual in ['MC', 'ASO', 'IDO', 'WAO', 'WD']:
                where = [h for h, value in enumerate(quals) if value == qual]
                for p in where:
                    if o_count > 5:
                        item = QTableWidgetItem(str(schedule[i][p][2]))
                        item.setTextAlignment(Qt.AlignCenter)
                        self.flight_schedule_table.setItem(3 * i + 2, 2 + 2 * (p - 6), item)
                        item = QTableWidgetItem(str(schedule[i][p][1]))
                        item.setTextAlignment(Qt.AlignCenter)
                        self.flight_schedule_table.setItem(3 * i + 2, 3 + 2 * (p - 6), item)
                        o_count += 1
                    elif o_count > 2:
                        item = QTableWidgetItem(str(schedule[i][p][2]))
                        item.setTextAlignment(Qt.AlignCenter)
                        self.flight_schedule_table.setItem(3 * i + 1, 2 + 2 * (p-3), item)
                        item = QTableWidgetItem(str(schedule[i][p][1]))
                        item.setTextAlignment(Qt.AlignCenter)
                        self.flight_schedule_table.setItem(3 * i + 1, 3 + 2 * (p-3), item)
                        o_count += 1
                    else:
                        item = QTableWidgetItem(str(schedule[i][p][2]))
                        item.setTextAlignment(Qt.AlignCenter)
                        self.flight_schedule_table.setItem(3*i, 2 +2*p, item)
                        item = QTableWidgetItem(str(schedule[i][p][1]))
                        item.setTextAlignment(Qt.AlignCenter)
                        self.flight_schedule_table.setItem(3*i, 3 + 2*p, item)
                        o_count += 1
            nco_count = 0
            for qual in ['SO1', 'SO2', 'SO3']:
                where = [h for h, value in enumerate(quals) if value == qual]
                for p in where:
                    if nco_count > 5:
                        item = QTableWidgetItem(str(schedule[i][p][2]))
                        item.setTextAlignment(Qt.AlignCenter)
                        self.flight_schedule_table.setItem(3 * i + 2, 8 + 2 * (p - 6 - o_count), item)
                        item = QTableWidgetItem(str(schedule[i][p][1]))
                        item.setTextAlignment(Qt.AlignCenter)
                        self.flight_schedule_table.setItem(3 * i + 2, 9 + 2 * (p - 6 - o_count), item)
                        nco_count += 1
                    elif nco_count > 2:
                        item = QTableWidgetItem(str(schedule[i][p][2]))
                        item.setTextAlignment(Qt.AlignCenter)
                        self.flight_schedule_table.setItem(3 * i + 1, 8 + 2 * (p - 3 - o_count), item)
                        item = QTableWidgetItem(str(schedule[i][p][1]))
                        item.setTextAlignment(Qt.AlignCenter)
                        self.flight_schedule_table.setItem(3 * i + 1, 9 + 2 * (p - 3 - o_count), item)
                        nco_count += 1
                    else:
                        item = QTableWidgetItem(str(schedule[i][p][2]))
                        item.setTextAlignment(Qt.AlignCenter)
                        self.flight_schedule_table.setItem(3 * i, 8 + 2 * (p-o_count), item)
                        item = QTableWidgetItem(str(schedule[i][p][1]))
                        item.setTextAlignment(Qt.AlignCenter)
                        self.flight_schedule_table.setItem(3 * i, 9 + 2 * (p-o_count), item)
                        nco_count += 1
        self.check_mode = 1

    def load_mission_info(self):
        mission_info = openpyxl.load_workbook('crew_data.xlsx')
        mission_info_sheet = mission_info.worksheets[2]

        index = 2
        mission = []
        while mission_info_sheet.cell(row=index, column=1).value != None:
            new_mission = []
            for i in range(1, 11):
                new_mission.append(str(mission_info_sheet.cell(row=index, column=i).value))
            for i in range(11, 22):
                if mission_info_sheet.cell(row=index, column=i).value != None:
                    new_mission.append((str(mission_info_sheet.cell(row=index, column=i).value)).split(','))
                else:
                    new_mission.append(['-'])
            mission.append(new_mission)
            index += 1

        self.mission_info_table.setRowCount(len(mission))
        for i in range(0, len(mission)):
            for k in range(0, 10):
                item = QTableWidgetItem(str(mission[i][k]))
                item.setTextAlignment(Qt.AlignCenter)
                self.mission_info_table.setItem(i, k, item)
            for k in range(10, len(mission[i])):
                if mission[i][k] != ['-']:
                    names = ''
                    if len(mission[i][k]) == 1:
                        names = str(mission[i][k][0])
                        item = QTableWidgetItem(names)
                        item.setTextAlignment(Qt.AlignCenter)
                    else:
                        names = str(mission[i][k][0])
                        for j in range(1, len(mission[i][k])):
                            names = names + ',' + str(mission[i][k][j])
                            item = QTableWidgetItem(names)
                            item.setTextAlignment(Qt.AlignCenter)
                    self.mission_info_table.setItem(i, k, item)
                else:
                    item = QTableWidgetItem('')
                    item.setTextAlignment(Qt.AlignCenter)
                    self.mission_info_table.setItem(i, k, item)

    def save_mission_info(self):
        old_crew_data = openpyxl.load_workbook('crew_data.xlsx')
        mission_info_sheet = old_crew_data['mission_info']
        for k in [2 for _ in range(100)]:
            mission_info_sheet.delete_rows(k)

        mission_info = []
        for i in range(0, self.mission_info_table.rowCount()):
            if self.mission_info_table.item(i, 0) != None:
                one_mission = []
                for k in range(0, 10):
                    if self.mission_info_table.item(i, k) != None:
                        one_mission.append(self.mission_info_table.item(i, k).text())
                    else:
                        one_mission.append(['-'])
                for k in range(10, 21):
                    if self.mission_info_table.item(i, k).text() != '':
                        one_mission.append((self.mission_info_table.item(i, k).text()).split(','))
                    else:
                        one_mission.append(['-'])
                mission_info.append(one_mission)

        for i in range(0, len(mission_info)):
            for j in range(0, 10):
                mission_info_sheet.cell(row=i+2, column=j+1, value=str(mission_info[i][j]))
            for j in range(10, len(mission_info[i])):
                if mission_info[i][j] != ['-']:
                    if len(mission_info[i][j]) == 1:
                        mission_info_sheet.cell(row=i+2, column=j+1, value=str(mission_info[i][j][0]))
                    else:
                        names = str(mission_info[i][j][0])
                        for k in range(1, len(mission_info[i][j])):
                            names = names + ',' + str(mission_info[i][j][k])
                        mission_info_sheet.cell(row=i + 2, column=j + 1, value=names)
        old_crew_data.save('crew_data.xlsx')

    def check_schedule(self, item):
        if self.check_mode == 1 and (item.column() == 3 or item.column() == 5 or item.column() == 7 or item.column() == 9 or item.column() == 11 or item.column() == 13):
            mission_num = item.row()//3 + 1
            rank = []
            if (item.column()-2)//6 == 0:
                rank = [3, 5, 7]
            elif (item.column()-2)//6 == 1:
                rank = [9, 11, 13]
            names = []
            if mission_num != 1:
                for i in rank:
                    for j in range(3*(mission_num - 2), 3*(mission_num - 1)):
                        if self.flight_schedule_table.item(j, i) != None:
                            if self.flight_schedule_table.item(j, i).text() != '':
                                names.append(self.flight_schedule_table.item(j, i).text())
                            else:
                                names.append('-')
                        else:
                            names.append('-')
            for i in rank:
                for j in range(3 * (mission_num - 1), 3 * (mission_num)):
                    if self.flight_schedule_table.item(j, i) != None:
                        if self.flight_schedule_table.item(j, i).text() != '':
                            names.append(self.flight_schedule_table.item(j, i).text())
                        else:
                            names.append('-')
                    else:
                        names.append('-')
            if mission_num != (self.flight_schedule_table.rowCount()) // 3:
                for i in rank:
                    for j in range(3 * (mission_num ), 3 * (mission_num + 1)):
                        if self.flight_schedule_table.item(j, i) != None:
                            if self.flight_schedule_table.item(j, i).text() != '':
                                names.append(self.flight_schedule_table.item(j, i).text())
                            else:
                                names.append('-')
                        else:
                            names.append('-')
            index = [i for i, value in enumerate(names) if value == item.text()]

            ex_names = []
            for i in range(14, 17):
                if self.flight_schedule_table.item(3*((item.row())//3), i) != None:
                    if self.flight_schedule_table.item(3*((item.row())//3), i).text() != '':
                        if len((self.flight_schedule_table.item(3*((item.row())//3), i).text()).split(',')) != 1:
                            for i in (self.flight_schedule_table.item(3*((item.row())//3), i).text()).split(','):
                                ex_names.append(i.replace(' ', ''))
                        else:
                            ex_names.append(self.flight_schedule_table.item(3*(item.row()//3), i).text())

            if len(index) != 1 and item.text() != '':
                item.setBackground(QtGui.QColor(255, 0, 0, 100))
                self.flight_schedule_table.item(item.row(), item.column() -1).setBackground(QtGui.QColor(255, 0, 0, 100))
                self.error_code.setText('    한 사람이 두번 연속의 소티에 포함될 수 없습니다!')
            elif item.text() in ex_names:
                item.setBackground(QtGui.QColor(255, 0, 0, 100))
                self.flight_schedule_table.item(item.row(), item.column() - 1).setBackground(
                    QtGui.QColor(255, 0, 0, 100))
                self.error_code.setText('    해당 인원은 해당 소티의 결원입니다')
            else:
                item.setBackground(QtGui.QColor(255, 255, 255))
                if self.flight_schedule_table.item(item.row(), item.column() - 1) != None:
                    self.flight_schedule_table.item(item.row(), item.column() - 1).setBackground(QtGui.QColor(255, 255, 255))
                self.error_code.setText('    문제 없음')